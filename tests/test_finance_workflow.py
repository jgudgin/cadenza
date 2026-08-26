"""The whole finance DAG, driven through the real orchestrator against
real Postgres, with only the two LLM call sites replaced by canned
responses - everything else (dependency resolution, the fan-out/fan-in
batch from validate_model's planner, the deterministic statement maths,
the actual Excel file) is exercised for real.

Deliberately not hitting the live API: this test should be fast,
deterministic, and runnable in CI with no ANTHROPIC_API_KEY - the same
principle as cadence's `cadence report`/`cadence check` never needing one.
A live run is a separate, manual thing (`cadenza run "Some Company"`).
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select

from cadenza.agents import registry
from cadenza.agents.finance import assumptions as assumptions_mod
from cadenza.agents.finance import summary as summary_mod
from cadenza.agents.finance import validate as validate_mod
from cadenza.models import Task, WorkflowRun
from cadenza.orchestrator import run_to_completion, start_run
from cadenza.registry import PlanOutcome, TaskSpec

GOOD_ASSUMPTIONS = {
    "starting_revenue": 50_000_000,
    "revenue_growth": 0.15,
    "gross_margin": 0.6,
    "opex_pct_revenue": 0.35,
    "da_pct_revenue": 0.04,
    "capex_pct_revenue": 0.05,
    "nwc_pct_revenue": 0.1,
    "tax_rate": 0.21,
    "starting_cash": 10_000_000,
    "starting_debt": 5_000_000,
    "rationale": "canned for testing",
}

EXPECTED_TASK_TYPES = {
    "gather_assumptions",
    "build_income_statement",
    "build_cash_flow",
    "build_balance_sheet",
    "validate_model",
    "sensitivity_analysis",
    "write_excel",
    "summarize_for_user",
}


async def _fake_assumptions(*, system, prompt, tool):
    return dict(GOOD_ASSUMPTIONS)


async def _fake_summary(*, system, prompt, tool):
    return {"summary": "This is a canned test summary."}


async def _fake_decide_next_steps(*, system, prompt):
    return PlanOutcome(
        reasoning="model balances and assumptions are plausible - fan out to sensitivity, then export",
        tasks=[
            TaskSpec(type="sensitivity_analysis", input={"scenario": "bull", "growth_delta": 0.03}, key="bull"),
            TaskSpec(type="sensitivity_analysis", input={"scenario": "base", "growth_delta": 0.0}, key="base"),
            TaskSpec(type="sensitivity_analysis", input={"scenario": "bear", "growth_delta": -0.03}, key="bear"),
            TaskSpec(type="write_excel", depends_on=["bull", "base", "bear"]),
        ],
    )


async def test_full_finance_pipeline_end_to_end(session_factory, monkeypatch, tmp_path):
    monkeypatch.setattr(assumptions_mod, "ask_claude_json", _fake_assumptions)
    monkeypatch.setattr(summary_mod, "ask_claude_json", _fake_summary)
    monkeypatch.setattr(validate_mod, "decide_next_steps", _fake_decide_next_steps)
    monkeypatch.chdir(tmp_path)

    run_id = await start_run(
        session_factory, "test model", TaskSpec(type="gather_assumptions", input={"company": "Acme", "round": 1})
    )
    await run_to_completion(session_factory, registry, run_id, concurrency=3, poll_interval=0.05)

    async with session_factory() as session:
        run = await session.get(WorkflowRun, run_id)
        tasks = (await session.execute(select(Task).where(Task.run_id == run_id))).scalars().all()

    assert run.status == "completed"
    assert all(t.status == "completed" for t in tasks), [(t.type, t.status, t.last_error) for t in tasks]
    assert {t.type for t in tasks} == EXPECTED_TASK_TYPES
    # Exactly one of each, except the three concurrent sensitivity scenarios.
    assert sum(1 for t in tasks if t.type == "sensitivity_analysis") == 3

    excel_path = Path(run.context["excel_path"])
    assert excel_path.exists()
    wb = load_workbook(excel_path)
    assert set(wb.sheetnames) == {
        "Assumptions",
        "Income Statement",
        "Cash Flow",
        "Balance Sheet",
        "Sensitivity",
    }
    assert run.context["summary"] == "This is a canned test summary."

    balance_sheet = run.context["balance_sheet"]
    for year_data in balance_sheet.values():
        assert abs(year_data["imbalance"]) < 1.0
