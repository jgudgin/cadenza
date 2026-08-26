"""The 'established Excel-based interface' end of the pipeline. Everything
upstream exists to produce numbers this step can lay out into a workbook -
this is deliberately the only place `openpyxl` is imported, so the rest of
the engine has no idea Excel is even involved.

openpyxl is synchronous, blocking I/O; it runs in a worker thread via
`asyncio.to_thread` rather than inline, so building a workbook never stalls
whatever else is happening concurrently on this worker (other tasks in
this run, or other runs entirely, sharing the same event loop).
"""

from __future__ import annotations

import asyncio
import os

from openpyxl import Workbook

from .. import registry
from ...exceptions import Permanent
from ...registry import AgentContext, PlanOutcome, TaskSpec

OUTPUT_DIR = "output"


def _slug(company: str) -> str:
    return "".join(c if c.isalnum() else "_" for c in company.strip().lower()) or "model"


def _write_statement(wb: Workbook, name: str, data: dict) -> None:
    sheet = wb.create_sheet(name)
    years = sorted(data.keys())
    row_labels = list(next(iter(data.values())).keys())
    sheet.append(["Line"] + years)
    for label in row_labels:
        row: list = [label]
        for year in years:
            value = data[year][label]
            row.append(round(value, 2) if isinstance(value, (int, float)) else value)
        sheet.append(row)


def _build_workbook(*, run_id: int, company: str, assumptions: dict, income_statement: dict,
                     cash_flow: dict, balance_sheet: dict, sensitivity: dict) -> str:
    wb = Workbook()
    assumptions_sheet = wb.active
    assumptions_sheet.title = "Assumptions"
    assumptions_sheet.append(["Assumption", "Value"])
    for key, value in assumptions.items():
        assumptions_sheet.append([key, value])

    _write_statement(wb, "Income Statement", income_statement)
    _write_statement(wb, "Cash Flow", cash_flow)
    _write_statement(wb, "Balance Sheet", balance_sheet)

    sens_sheet = wb.create_sheet("Sensitivity")
    sens_sheet.append(["Scenario", "Growth Delta", "Final Year Revenue", "Final Year Net Income", "Final Year Cash"])
    for scenario in ("bull", "base", "bear"):
        s = sensitivity[scenario]
        sens_sheet.append(
            [
                s["scenario"],
                s["growth_delta"],
                round(s["final_year_revenue"], 2),
                round(s["final_year_net_income"], 2),
                round(s["final_year_cash"], 2),
            ]
        )

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    path = os.path.join(OUTPUT_DIR, f"{_slug(company)}-run{run_id}.xlsx")
    wb.save(path)
    return path


async def write_excel(ctx: AgentContext) -> dict:
    rc = ctx.run_context
    try:
        company = rc["company"]
        assumptions = rc["assumptions"]
        income_statement = rc["income_statement"]
        cash_flow = rc["cash_flow_statement"]
        balance_sheet = rc["balance_sheet"]
        sensitivity = {s: rc[f"sensitivity_{s}"] for s in ("bull", "base", "bear")}
    except KeyError as exc:
        raise Permanent(f"missing required context: {exc}") from exc

    path = await asyncio.to_thread(
        _build_workbook,
        run_id=ctx.run_id,
        company=company,
        assumptions=assumptions,
        income_statement=income_statement,
        cash_flow=cash_flow,
        balance_sheet=balance_sheet,
        sensitivity=sensitivity,
    )
    return {"excel_path": path}


async def plan_next_write_excel(input) -> PlanOutcome:  # noqa: ANN001
    return PlanOutcome(
        reasoning="Workbook written; producing a narrative summary to close out the run.",
        context_updates={"excel_path": input.output["excel_path"]},
        tasks=[TaskSpec(type="summarize_for_user", depends_on=[input.task_id])],
    )


registry.agent("write_excel", plan_next=plan_next_write_excel)(write_excel)
